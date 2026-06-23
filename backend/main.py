from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
from dotenv import load_dotenv
from supabase import create_client
import aiosmtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from twilio.rest import Client as TwilioClient
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
import openpyxl
import shutil
import io
import os
import base64
import resend

load_dotenv()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Supabase
supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))

# Twilio
twilio = TwilioClient(os.getenv("TWILIO_ACCOUNT_SID"), os.getenv("TWILIO_AUTH_TOKEN"))

# Resend
resend.api_key = os.getenv("RESEND_API_KEY")
print("RESEND KEY LOADED:", repr(resend.api_key))
SENDER_EMAIL = os.getenv("SENDER_EMAIL", "saurav@bharyat.com")
SIR_EMAIL = os.getenv("SIR_EMAIL", "saurav@bharyat.com")

FRONTEND_BASE_URL = os.getenv("FRONTEND_BASE_URL", "http://127.0.0.1:5500")

# ---------- Models ----------

class Customer(BaseModel):
    name: str
    email: Optional[str] = None
    whatsapp: Optional[str] = None

class Recipient(BaseModel):
    name: str
    email: Optional[str] = None
    whatsapp: Optional[str] = None

class SendMessageRequest(BaseModel):
    message: str
    subject: Optional[str] = "Message from Bharyat"
    recipients: List[Recipient]

class RFQResponseItem(BaseModel):
    part_number: str
    data_code: Optional[str] = ""
    unit_price: Optional[str] = ""
    delivery_time: Optional[str] = ""
    notes: Optional[str] = ""

class RFQSubmitRequest(BaseModel):
    responses: List[RFQResponseItem]

# ---------- Customers ----------

@app.get("/api/customers")
def get_customers():
    res = supabase.table("customers").select("*").order("created_at", desc=True).execute()
    return res.data

@app.post("/api/customers")
def add_customer(c: Customer):
    if not c.email and not c.whatsapp:
        raise HTTPException(400, "Email or WhatsApp required")
    res = supabase.table("customers").insert({
        "name": c.name,
        "email": c.email,
        "whatsapp": c.whatsapp
    }).execute()
    return res.data[0]

@app.delete("/api/customers/{customer_id}")
def delete_customer(customer_id: int):
    supabase.table("customers").delete().eq("id", customer_id).execute()
    return {"success": True}

# ---------- Send Message (via Resend) ----------

def send_email(to_email: str, subject: str, body: str, reply_to: Optional[str] = None, attachment_bytes: Optional[bytes] = None, attachment_filename: Optional[str] = None):
    params = {
        "from": f"Bharyat Advanced Systems <{SENDER_EMAIL}>",
        "to": [to_email],
        "subject": subject,
        "text": body,
    }
    if reply_to:
        params["reply_to"] = reply_to
    if attachment_bytes and attachment_filename:
        encoded = base64.b64encode(attachment_bytes).decode("utf-8")
        params["attachments"] = [{
            "filename": attachment_filename,
            "content": encoded,
        }]
    resend.Emails.send(params)

def send_whatsapp(to_number: str, body: str):
    if not to_number.startswith('+'):
        to_number = '+' + to_number
    twilio.messages.create(
        from_=os.getenv("TWILIO_WHATSAPP_FROM"),
        to=f"whatsapp:{to_number}",
        body=body
    )

@app.post("/api/send-message")
async def send_message(req: SendMessageRequest):
    results = {"email": [], "whatsapp": [], "errors": []}

    for r in req.recipients:
        if r.email:
            try:
                send_email(r.email, req.subject, req.message)
                results["email"].append(r.name)
                supabase.table("message_logs").insert({
                    "customer_name": r.name,
                    "channel": "email",
                    "subject": req.subject,
                    "message": req.message,
                    "status": "sent"
                }).execute()
            except Exception as e:
                results["errors"].append({"name": r.name, "channel": "email", "error": str(e)})
                supabase.table("message_logs").insert({
                    "customer_name": r.name,
                    "channel": "email",
                    "subject": req.subject,
                    "message": req.message,
                    "status": "failed",
                    "error_text": str(e)
                }).execute()

        if r.whatsapp:
            try:
                send_whatsapp(r.whatsapp, req.message)
                results["whatsapp"].append(r.name)
                supabase.table("message_logs").insert({
                    "customer_name": r.name,
                    "channel": "whatsapp",
                    "message": req.message,
                    "status": "sent"
                }).execute()
            except Exception as e:
                results["errors"].append({"name": r.name, "channel": "whatsapp", "error": str(e)})
                supabase.table("message_logs").insert({
                    "customer_name": r.name,
                    "channel": "whatsapp",
                    "message": req.message,
                    "status": "failed",
                    "error_text": str(e)
                }).execute()

    return {"success": True, "results": results}

# ---------- Logs ----------

@app.get("/api/logs")
def get_logs():
    res = supabase.table("message_logs").select("*").order("sent_at", desc=True).limit(100).execute()
    return res.data

# ---------- Upload Customers Excel ----------

@app.post("/api/customers/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

    name_col = next((i for i, h in enumerate(headers) if 'name' in h), None)
    email_col = next((i for i, h in enumerate(headers) if 'email' in h or 'mail' in h), None)
    wa_col = next((i for i, h in enumerate(headers) if 'whatsapp' in h or 'mobile' in h or 'phone' in h or 'wa' in h), None)

    if name_col is None:
        raise HTTPException(400, "Name column not found in Excel")

    added = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[name_col]).strip() if row[name_col] else None
        email = str(row[email_col]).strip() if email_col is not None and row[email_col] else None
        wa = str(row[wa_col]).strip() if wa_col is not None and row[wa_col] else None

        if not name or name == 'None':
            skipped += 1
            continue
        if not email and not wa:
            skipped += 1
            continue

        res = supabase.table("customers").insert({
            "name": name,
            "email": email,
            "whatsapp": wa
        }).execute()
        added.append(res.data[0])

    return {"success": True, "added": len(added), "skipped": skipped, "customers": added}

# ---------- RFQ Send (with file upload) ----------

@app.post("/api/rfq/send")
async def send_rfq(
    file: UploadFile = File(...),
    supplier_ids: str = Form(...),
    subject: str = Form(...),
    message: str = Form(...)
):
    file_bytes = await file.read()

    ids = [int(i) for i in supplier_ids.split(",")]
    res = supabase.table("customers").select("*").in_("id", ids).execute()
    suppliers = res.data
    results = {"sent": [], "errors": []}

    for supplier in suppliers:
        if supplier.get("email"):
            try:
                full_message = message + f"\n\nPlease reply to this email with the filled Excel sheet.\nEmail: {SIR_EMAIL}"
                send_email(
                    supplier["email"],
                    subject,
                    full_message,
                    reply_to=SIR_EMAIL,
                    attachment_bytes=file_bytes,
                    attachment_filename=file.filename
                )
                results["sent"].append(supplier["name"])
                supabase.table("message_logs").insert({
                    "customer_name": supplier["name"],
                    "channel": "email",
                    "subject": subject,
                    "message": message,
                    "status": "sent"
                }).execute()
            except Exception as e:
                results["errors"].append({"name": supplier["name"], "error": str(e)})

    return {"success": True, "results": results}

# ---------- RFQ Generate + Send (web form link, not Excel) ----------

@app.post("/api/rfq/send-with-parts")
async def send_rfq_with_parts(
    supplier_ids: str = Form(...),
    subject: str = Form(...),
    message: str = Form(...),
    part_numbers: str = Form(...),
):
    ids = [int(i) for i in supplier_ids.split(",")]
    res = supabase.table("customers").select("*").in_("id", ids).execute()
    suppliers = res.data
    results = {"sent": [], "errors": []}

    clean_parts = ",".join([p.strip() for p in part_numbers.split(",") if p.strip()])

    for supplier in suppliers:
        if supplier.get("email"):
            try:
                # Create one RFQ batch record per supplier
                batch_res = supabase.table("rfq_batches").insert({
                    "supplier_id": supplier["id"],
                    "supplier_name": supplier["name"],
                    "part_numbers": clean_parts,
                    "subject": subject,
                    "message": message,
                    "status": "pending"
                }).execute()
                rfq_id = batch_res.data[0]["id"]

                form_link = f"{FRONTEND_BASE_URL}/rfq-form.html?id={rfq_id}"
                full_message = message + f"\n\nPlease fill in your quotation using this secure form:\n{form_link}\n\nIf you have questions, reply to this email."

                send_email(
                    supplier["email"],
                    subject,
                    full_message,
                    reply_to=SIR_EMAIL,
                )
                results["sent"].append(supplier["name"])
                supabase.table("message_logs").insert({
                    "customer_name": supplier["name"],
                    "channel": "email",
                    "subject": subject,
                    "message": message,
                    "status": "sent"
                }).execute()

            except Exception as e:
                print(f"ERROR sending to {supplier['name']}: {e}")
                results["errors"].append({"name": supplier["name"], "error": str(e)})

    return {"success": True, "results": results}

# ---------- RFQ Form: fetch details ----------

@app.get("/api/rfq/{rfq_id}")
def get_rfq(rfq_id: str):
    res = supabase.table("rfq_batches").select("*").eq("id", rfq_id).execute()
    if not res.data:
        raise HTTPException(404, "RFQ not found")
    return res.data[0]

# ---------- RFQ Form: submit supplier response ----------

@app.post("/api/rfq/{rfq_id}/submit")
def submit_rfq(rfq_id: str, req: RFQSubmitRequest):
    batch_res = supabase.table("rfq_batches").select("*").eq("id", rfq_id).execute()
    if not batch_res.data:
        raise HTTPException(404, "RFQ not found")
    batch = batch_res.data[0]
    if batch["status"] == "submitted":
        raise HTTPException(400, "This RFQ has already been submitted")

    for item in req.responses:
        supabase.table("rfq_responses").insert({
            "rfq_batch_id": rfq_id,
            "part_number": item.part_number,
            "data_code": item.data_code,
            "unit_price": item.unit_price,
            "delivery_time": item.delivery_time,
            "notes": item.notes,
        }).execute()

    supabase.table("rfq_batches").update({"status": "submitted"}).eq("id", rfq_id).execute()

    # Notify Sir via email
    try:
        rows_text = "\n".join([
            f"- Part: {item.part_number} | Data Code: {item.data_code or '-'} | Price: {item.unit_price or '-'} | Delivery: {item.delivery_time or '-'} | Notes: {item.notes or '-'}"
            for item in req.responses
        ])
        notify_body = (
            f"Supplier {batch['supplier_name']} has submitted their RFQ response.\n\n"
            f"Subject: {batch['subject']}\n\n"
            f"Details:\n{rows_text}\n"
        )
        send_email(
            SIR_EMAIL,
            f"RFQ Response Received from {batch['supplier_name']}",
            notify_body,
        )
    except Exception as e:
        print(f"Failed to notify Sir: {e}")

    return {"success": True}

# ---------- RFQ Responses (for dashboard) ----------

@app.get("/api/rfq-responses")
def get_rfq_responses():
    batches = supabase.table("rfq_batches").select("*").order("created_at", desc=True).execute().data
    responses = supabase.table("rfq_responses").select("*").order("submitted_at", desc=True).execute().data
    return {"batches": batches, "responses": responses}